from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pyautogui
import time
import random
import pyperclip
import subprocess
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

# ------------------------------------------------------------------------------------

# Lista de CNPJs para processamento
cnpj_list = ["00000000000000"] #Aqui voce coloca os CNPJs que quer consultar

# ------------------------------------------------------------------------------------

def espera_aleatoria():
    time.sleep(random.uniform(1.5, 5))
    
def abrir_chrome_modo_depuracao_com_site(porta=9222, url="https://cav.receita.fazenda.gov.br/autenticacao/login"):
    # Caminho padrão para o Google Chrome, ajuste se necessário
    caminho_chrome = "Caminho/do/chrome.exe"
    
    # Parâmetros para abrir o Chrome em modo de depuração
    comando = [
        caminho_chrome,
        f"--remote-debugging-port={porta}",
        "--user-data-dir=C:\\chrome_debug_profile",
        url  # Adiciona a URL ao comando
    ]
    
    try:
        subprocess.Popen(comando)
        print(f"Chrome aberto em modo de depuração na porta {porta}. Site: {url}")
    except FileNotFoundError:
        print("Erro: Caminho do Google Chrome não encontrado. Verifique o caminho no código.")
    except Exception as e:
        print(f"Ocorreu um erro ao abrir o Chrome: {e}")

# ------------------------------------------------------------------------------------

def clicar_botao(x, y, descricao):
    espera_aleatoria()
    # Move o mouse para o botão e clica
    pyautogui.moveTo(x, y, duration=1)
    pyautogui.click()
    print(f"Botão '{descricao}' clicado com sucesso!")

def EntrarcomGov():
    espera_aleatoria()
    # Coordenadas do botão "Entrar com Gov"
    x = 885  # Substitua pelo valor de X encontrado
    y = 479  # Substitua pelo valor de Y encontrado
    espera_aleatoria()
    clicar_botao(x, y, "Entrar com Gov")

def SeuCertificadoDigital():
    espera_aleatoria()
    # Coordenadas do botão "Seu Certificado Digital"
    x = 945  # Substitua pelo valor de X encontrado
    y = 647  # Substitua pelo valor de Y encontrado
    espera_aleatoria()  # Aguarda a interface carregar antes de clicar
    clicar_botao(x, y, "Seu Certificado Digital")
    
def OkCertificado():
    espera_aleatoria()
    # Coordenadas do botão "Seu Certificado Digital"
    x = 799  # Substitua pelo valor de X encontrado
    y = 357  # Substitua pelo valor de Y encontrado
    espera_aleatoria()  # Aguarda a interface carregar antes de clicar
    clicar_botao(x, y, "OkCertificado")
    
def AlterarPerfil():
    espera_aleatoria()
    # Coordenadas do botão "Seu Certificado Digital"
    x = 1080  # Substitua pelo valor de X encontrado
    y = 219  # Substitua pelo valor de Y encontrado
    espera_aleatoria()  # Aguarda a interface carregar antes de clicar
    clicar_botao(x, y, "AlterarPerfil")
    
def ClicarCNPJ():
    espera_aleatoria()
    # Coordenadas do botão "Seu Certificado Digital"
    x = 542  # Substitua pelo valor de X encontrado
    y = 381  # Substitua pelo valor de Y encontrado
    espera_aleatoria()  # Aguarda a interface carregar antes de clicar
    clicar_botao(x, y, "ClicarCNPJ")

def ColarCNPJ():
    espera_aleatoria()
    try:
        # Verifica se há CNPJs na lista
        if not cnpj_list:
            print("Nenhum CNPJ disponível na lista para colar.")
            return

        # Copiar o primeiro CNPJ da lista para a área de transferência
        cnpj = cnpj_list.pop(0)  # Remove e retorna o primeiro CNPJ
        pyperclip.copy(cnpj)

        # Simula colagem do CNPJ no campo (Ctrl + V)
        espera_aleatoria()
        pyautogui.hotkey("ctrl", "v")
        espera_aleatoria()

        print(f"CNPJ {cnpj} colado com sucesso!")
    except Exception as e:
        print(f"Erro em ColarCNPJ: {e}")
        
def ClicarAlterar():
    espera_aleatoria()  # Aguarda um tempo aleatório para parecer mais humano
    try:
        # Simula o pressionamento da tecla Tab para navegar
        pyautogui.press("tab")
        espera_aleatoria()

        # Simula o pressionamento da tecla Enter para ativar o botão
        pyautogui.press("enter")
        espera_aleatoria()

        print("Ação 'ClicarAlterar' realizada com sucesso usando Tab e Enter!")
    
    except Exception as e:
        print(f"Erro ao executar 'ClicarAlterar': {e}")
        
def FecharJanela():
    try:
        # Limpar o cache do navegador via teclas de atalho (Ctrl + Shift + Del)
        espera_aleatoria()
        pyautogui.hotkey('ctrl', 'shift', 'del')
        espera_aleatoria()
        
        # Aguarda a interface abrir e pressiona Enter para confirmar
        pyautogui.press('enter')
        espera_aleatoria()

        # Fecha o navegador
        pyautogui.hotkey('alt', 'f4')
        print("Cache limpo e navegador fechado com sucesso.")
    except Exception as e:
        print(f"Erro ao tentar limpar cache e fechar navegador: {e}")
    
# ------------------------------------------------------------------------------------
    
def NovasMensagens():
    try:
        # Localizar o iframe, se necessário
        #iframe = driver.find_element(By.TAG_NAME, "iframe")  # Substitua por um localizador específico do iframe, se necessário
        #driver.switch_to.frame(iframe)

        # Aguarde até que o elemento com o texto "RECEITA FEDERAL DO BRASIL" esteja visível
        wait = WebDriverWait(driver, 10)  # Aguarde até 10 segundos
        elemento = wait.until(EC.presence_of_element_located((By.ID, "btnCaixaPostal")))

        # Clicar no elemento
        elemento.click()
        print("NovasMensagens clicado com sucesso!")

        # Voltar para o contexto principal
        #driver.switch_to.default_content()

    except Exception as e:
        print("Erro encontrado:", e)

    finally:
        # Não fechar o navegador automaticamente para continuar na aba
        pass
    
def ReceitaFederal():
    try:
        # Localizar o iframe, se necessário
        iframe = driver.find_element(By.TAG_NAME, "iframe")  # Substitua por um localizador específico do iframe, se necessário
        driver.switch_to.frame(iframe)

        # Aguarde até que o elemento com o atributo src="imagens/aaMsgNaoLida.gif" esteja visível
        wait = WebDriverWait(driver, 10)  # Aguarde até 10 segundos
        elemento = wait.until(EC.presence_of_element_located((By.XPATH, "//img[@src='imagens/aaMsgNaoLida.gif']")))
        # Localizar o elemento "Receita Federal do Brasil" na mesma linha do elemento encontrado
        receita = elemento.find_element(By.XPATH, "./ancestor::tr//*[contains(text(), 'RECEITA FEDERAL DO BRASIL')]")

        # Clicar no elemento "Receita Federal do Brasil"
        receita.click()
        print("Elemento 'RECEITA FEDERAL DO BRASIL' clicado com sucesso!")

        # Voltar para o contexto principal
        driver.switch_to.default_content()
        return True  # Indica que encontrou a mensagem
    except Exception as e:
        print("Nenhuma mensagem não lida encontrada ou erro em ReceitaFederal:", e)
        driver.switch_to.default_content()
        return False  # Indica que não encontrou a mensagem

    except Exception as e:
        print("Erro encontrado:", e)

    finally:
        # Não fechar o navegador automaticamente para continuar na aba
        pass
    
def salvar_mensagens_no_excel(df, arquivo_excel):
    try:
        # Verifica se o arquivo já existe
        try:
            # Abre o arquivo existente
            workbook = load_workbook(arquivo_excel)
            sheet = workbook.active
        except FileNotFoundError:
            # Cria um novo arquivo se não existir
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = 'Mensagens'
            # Adiciona cabeçalhos na primeira linha
            headers = ["Assunto:", "Enviada em:", "Primeira Leitura:", "Exibição até:", "CNPJ do destinatário:", "Mensagem:"]
            sheet.append(headers)

        # Adiciona os novos dados na próxima linha disponível
        for row in dataframe_to_rows(df, index=False, header=False):
            sheet.append(row)

        # Ajusta a largura das colunas
        ajustar_largura_colunas(sheet)

        # Salva o arquivo
        workbook.save(arquivo_excel)
        print(f"Mensagens salvas no Excel com sucesso! Arquivo salvo em '{arquivo_excel}'.")

    except Exception as e:
        print(f"Erro ao salvar mensagens no Excel: {e}")

def ajustar_largura_colunas(sheet):
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter  # Obtém a letra da coluna
        for cell in column:
            try:  # Calcula o comprimento máximo do conteúdo
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2  # Adiciona um pequeno buffer
        sheet.column_dimensions[column_letter].width = adjusted_width

def CopiarMensagem():
    global df  # Permite modificar o DataFrame global
    try:
        # Localizar o iframe, se necessário
        iframe = driver.find_element(By.TAG_NAME, "iframe")  # Substitua por um localizador específico do iframe, se necessário
        driver.switch_to.frame(iframe)

        # Aguarde até que os elementos estejam visíveis e copie os conteúdos
        wait = WebDriverWait(driver, 10)  # Aguarde até 10 segundos
        
        # Copiando os valores dos IDs fornecidos
        assunto = wait.until(EC.presence_of_element_located((By.ID, "assunto"))).text
        enviada_em = wait.until(EC.presence_of_element_located((By.ID, "dtEnvio"))).text
        primeira_leitura = wait.until(EC.presence_of_element_located((By.ID, "lbValorPrimeiraLeitura"))).text
        exibicao_ate = wait.until(EC.presence_of_element_located((By.ID, "dtExpiracao"))).text
        cnpj_destinatario = wait.until(EC.presence_of_element_located((By.ID, "lbValorCNPJReferencia"))).text
        mensagem = wait.until(EC.presence_of_element_located((By.ID, "msgConteudo"))).text

        # Adicionar os dados ao DataFrame
        df = pd.DataFrame([{
            "Assunto:": assunto,
            "Enviada em:": enviada_em,
            "Primeira Leitura:": primeira_leitura,
            "Exibição até:": exibicao_ate,
            "CNPJ do destinatário:": cnpj_destinatario,
            "Mensagem:": mensagem
        }])

        # Salvar no Excel imediatamente após copiar
        arquivo_excel = "Caminho/para/salvar/o/arquivo/excel"
        salvar_mensagens_no_excel(df, arquivo_excel)

        # Voltar para o contexto principal
        driver.switch_to.default_content()

        print("Mensagem copiada com sucesso!")

    except Exception as e:
        print("Erro encontrado:", e)

    finally:
        # Não fechar o navegador automaticamente para continuar na aba
        pass
        
def TelaInicial():
    try:
        # Localizar o iframe, se necessário
        #iframe = driver.find_element(By.TAG_NAME, "iframe")  # Substitua por um localizador específico do iframe, se necessário
        #driver.switch_to.frame(iframe)

        # Aguarde até que o elemento com o texto "RECEITA FEDERAL DO BRASIL" esteja visível
        wait = WebDriverWait(driver, 10)  # Aguarde até 10 segundos
        elemento = wait.until(EC.presence_of_element_located((By.ID, "logoEcac")))

        # Clicar no elemento
        elemento.click()
        print("TelaInicial clicado com sucesso!")

        # Voltar para o contexto principal
        #driver.switch_to.default_content()

    except Exception as e:
        print("Erro encontrado:", e)

    finally:
        # Não fechar o navegador automaticamente para continuar na aba
        pass

def MensagensImportantes():
    # Verifica se o botão com o texto "Ir para a Caixa Postal" existe na página
    if driver.find_elements(By.XPATH, "//button[text()='Ir para a Caixa Postal']"):
        print("Botão 'Ir para a Caixa Postal' encontrado. Interagindo com o elemento...")

        # Aguarde até que o botão esteja visível e clicável
        wait = WebDriverWait(driver, 10)  # Aguarde até 10 segundos
        elemento = wait.until(EC.presence_of_element_located((By.XPATH, "//button[text()='Ir para a Caixa Postal']")))

        # Clicar no botão
        elemento.click()
        print("Botão 'Ir para a Caixa Postal' clicado com sucesso!")
    else:
        print("Botão 'Ir para a Caixa Postal' não encontrado. Prosseguindo com o restante do código.")
        
# ------------------------------------------------------------------------------------

if __name__ == "__main__":
    # Loop para processar todos os CNPJs
    while cnpj_list:  # Continua enquanto houver CNPJs na lista
        try:
            # Executa as ações necessárias para um CNPJ
            abrir_chrome_modo_depuracao_com_site()
            EntrarcomGov()
            SeuCertificadoDigital()
            OkCertificado()
            AlterarPerfil()
            ClicarCNPJ()
            ColarCNPJ()
            ClicarAlterar()
            
            # Configuração do WebDriver em modo de depuração
            chrome_options = Options()
            chrome_options.add_experimental_option("debuggerAddress", "127.0.0.1:9222")  # Porta para depuração

            driver_path = 'Caminho/chromedriver.exe'  # Substitua pelo caminho do seu WebDriver
            service = Service(driver_path)

            # Conectando ao navegador já aberto
            driver = webdriver.Chrome(service=service, options=chrome_options)

            # Loop interno para processar mensagens para o CNPJ atual
            while True:
                MensagensImportantes()
                espera_aleatoria()
                NovasMensagens()
                espera_aleatoria()
                encontrou_mensagem = ReceitaFederal()
                if not encontrou_mensagem:  # Quando não há mais mensagens, saia do loop interno
                    print("Não há mais mensagens para o CNPJ atual.")
                    TelaInicial()
                    FecharJanela()
                    break
                espera_aleatoria()
                CopiarMensagem()
                espera_aleatoria()
            
            print("CNPJ processado com sucesso. Continuando para o próximo...")
        except Exception as e:
            print(f"Erro no processamento do CNPJ: {e}")
            break  # Encerra em caso de erro grave
        
    print("Todas as mensagens não lidas foram copiadas e salvas com sucesso!")     
       
# ------------------------------------------------------------------------------------
